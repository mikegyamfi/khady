import hashlib
import hmac
import json
from datetime import datetime
from time import sleep

import pandas as pd
from decouple import config
from django.contrib.auth.models import Group, Permission
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponseRedirect, HttpResponse
import requests
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from tablib import Dataset

from . import forms
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from . import helper, models
from .forms import UploadFileForm, OrderForm, PackageForm, TrackingForm, StatusUpdateForm
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
from .models import MTNTransaction, Tracking  # Adjust the import based on your model's location

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from openpyxl import load_workbook


# Create your views here.
def home(request):
    return render(request, "layouts/index.html")


def services(request):
    return render(request, "layouts/services.html")


@login_required(login_url='login')
def pay_with_wallet(request):
    if request.method == "POST":
        admin = models.AdminInfo.objects.filter().first().phone_number
        user = models.CustomUser.objects.get(id=request.user.id)
        phone_number = request.POST.get("phone")
        amount = request.POST.get("amount")
        reference = request.POST.get("reference")
        if user.wallet is None:
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge. Admin Contact Info: 0{admin}'})
        elif user.wallet <= 0 or user.wallet < float(amount):
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge. Admin Contact Info: 0{admin}'})
        print(phone_number)
        print(amount)
        print(reference)
        if user.status == "User":
            bundle = models.IshareBundlePrice.objects.get(price=float(amount)).bundle_volume
        elif user.status == "Agent":
            bundle = models.AgentIshareBundlePrice.objects.get(price=float(amount)).bundle_volume
        elif user.status == "Super Agent":
            bundle = models.SuperAgentIshareBundlePrice.objects.get(price=float(amount)).bundle_volume
        else:
            bundle = models.IshareBundlePrice.objects.get(price=float(amount)).bundle_volume

        # print(bundle)
        # send_bundle_response = helper.send_bundle(request.user, phone_number, bundle, reference)
        # data = send_bundle_response.json()
        # print(data)

        sms_headers = {
            'Authorization': 'Bearer 1334|wroIm5YnQD6hlZzd8POtLDXxl4vQodCZNorATYGX',
            'Content-Type': 'application/json'
        }

        sms_url = 'https://webapp.usmsgh.com/api/sms/send'

        send_bundle_response = helper.send_bundle(phone_number, bundle, reference)
        try:
            data = send_bundle_response.json()
            print(data)
        except:
            return JsonResponse({'status': f'Something went wrong'})

        sms_headers = {
            'Authorization': 'Bearer 1334|wroIm5YnQD6hlZzd8POtLDXxl4vQodCZNorATYGX',
            'Content-Type': 'application/json'
        }

        sms_url = 'https://webapp.usmsgh.com/api/sms/send'
        if send_bundle_response.status_code == 200:
            if data["status"] == "Success":
                new_transaction = models.IShareBundleTransaction.objects.create(
                    user=request.user,
                    bundle_number=phone_number,
                    offer=f"{bundle}MB",
                    reference=reference,
                    transaction_status="Completed"
                )
                new_transaction.save()
                user.wallet -= float(amount)
                user.save()
                receiver_message = f"Your bundle purchase has been completed successfully. {bundle}MB has been credited to you by {request.user.phone}.\nReference: {reference}\n"
                sms_message = f"Hello @{request.user.username}. Your bundle purchase has been completed successfully. {bundle}MB has been credited to {phone_number}.\nReference: {reference}\nCurrent Wallet Balance: {user.wallet}\nThank you for using GH BAY."

                num_without_0 = phone_number[1:]
                print(num_without_0)
                receiver_body = {
                    'recipient': f"233{num_without_0}",
                    'sender_id': 'GH BAY',
                    'message': receiver_message
                }

                response = requests.request('POST', url=sms_url, params=receiver_body, headers=sms_headers)
                print(response.text)

                sms_body = {
                    'recipient': f"233{request.user.phone}",
                    'sender_id': 'GH BAY',
                    'message': sms_message
                }

                response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)

                print(response.text)

                return JsonResponse({'status': 'Transaction Completed Successfully', 'icon': 'success'})
            else:
                new_transaction = models.IShareBundleTransaction.objects.create(
                    user=request.user,
                    bundle_number=phone_number,
                    offer=f"{bundle}MB",
                    reference=reference,
                    transaction_status="Failed"
                )
                new_transaction.save()
                return JsonResponse({'status': 'Something went wrong', 'icon': 'error'})
        else:
            new_transaction = models.IShareBundleTransaction.objects.create(
                user=request.user,
                bundle_number=phone_number,
                offer=f"{bundle}MB",
                reference=reference,
                transaction_status="Failed"
            )
            new_transaction.save()
            return JsonResponse({'status': 'Something went wrong', 'icon': 'error'})
    return redirect('airtel-tigo')


@login_required(login_url='login')
def airtel_tigo(request):
    user = models.CustomUser.objects.get(id=request.user.id)
    status = user.status
    form = forms.IShareBundleForm(status)
    reference = helper.ref_generator()
    db_user_id = request.user.id
    user_email = request.user.email
    if request.method == "POST":
        form = forms.IShareBundleForm(data=request.POST, status=status)
        payment_reference = request.POST.get("reference")
        amount_paid = request.POST.get("amount")
        new_payment = models.Payment.objects.create(
            user=request.user,
            reference=payment_reference,
            amount=amount_paid,
            transaction_date=datetime.now(),
            transaction_status="Completed"
        )
        new_payment.save()
        print("payment saved")
        print("form valid")
        phone_number = request.POST.get("phone")
        offer = request.POST.get("amount")
        print(offer)
        if user.status == "User":
            bundle = models.IshareBundlePrice.objects.get(price=float(offer)).bundle_volume
        elif user.status == "Agent":
            bundle = models.AgentIshareBundlePrice.objects.get(price=float(offer)).bundle_volume
        elif user.status == "Super Agent":
            bundle = models.SuperAgentIshareBundlePrice.objects.get(price=float(offer)).bundle_volume
        else:
            bundle = models.IshareBundlePrice.objects.get(price=float(offer)).bundle_volume
        new_transaction = models.IShareBundleTransaction.objects.create(
            user=request.user,
            bundle_number=phone_number,
            offer=f"{bundle}MB",
            reference=payment_reference,
            transaction_status="Pending"
        )
        print("created")
        new_transaction.save()

        print("===========================")
        print(phone_number)
        print(bundle)
        print("--------------------")
        # send_bundle_response = helper.send_bundle(request.user, phone_number, bundle, payment_reference)
        # print("********************")
        # data = send_bundle_response.json()
        #
        # print(data)
        return JsonResponse({'status': 'Transaction Completed Successfully', 'icon': 'success'})
    user = models.CustomUser.objects.get(id=request.user.id)
    context = {"form": form, "ref": reference, 'id': db_user_id, "email": user_email,
               "wallet": 0 if user.wallet is None else user.wallet}
    return render(request, "layouts/services/at.html", context=context)


@login_required(login_url='login')
def mtn_pay_with_wallet(request):
    if request.method == "POST":
        user = models.CustomUser.objects.get(id=request.user.id)
        phone_number = request.POST.get("phone")
        amount = request.POST.get("amount")
        reference = request.POST.get("reference")
        print(phone_number)
        print(amount)
        print(reference)
        if user.wallet is None:
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge.'})
        elif user.wallet <= 0 or user.wallet < float(amount):
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge.'})
        if user.status == "User":
            bundle = models.MTNBundlePrice.objects.get(price=float(amount)).bundle_volume
        elif user.status == "Agent":
            bundle = models.AgentMTNBundlePrice.objects.get(price=float(amount)).bundle_volume
        elif user.status == "Super Agent":
            bundle = models.SuperAgentMTNBundlePrice.objects.get(price=float(amount)).bundle_volume
        else:
            bundle = models.MTNBundlePrice.objects.get(price=float(amount)).bundle_volume
        print(bundle)
        new_mtn_transaction = models.MTNTransaction.objects.create(
            user=request.user,
            bundle_number=phone_number,
            offer=f"{bundle}MB",
            reference=reference,
        )
        new_mtn_transaction.save()
        user.wallet -= float(amount)
        user.save()
        return JsonResponse({'status': "Your transaction will be completed shortly", 'icon': 'success'})
    return redirect('mtn')


@login_required(login_url='login')
def big_time_pay_with_wallet(request):
    if request.method == "POST":
        user = models.CustomUser.objects.get(id=request.user.id)
        phone_number = request.POST.get("phone")
        amount = request.POST.get("amount")
        reference = request.POST.get("reference")
        print(phone_number)
        print(amount)
        print(reference)
        if user.wallet is None:
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge.'})
        elif user.wallet <= 0 or user.wallet < float(amount):
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge.'})
        if user.status == "User":
            bundle = models.BigTimeBundlePrice.objects.get(price=float(amount)).bundle_volume
        elif user.status == "Agent":
            bundle = models.AgentBigTimeBundlePrice.objects.get(price=float(amount)).bundle_volume
        elif user.status == "Super Agent":
            bundle = models.SuperAgentBigTimeBundlePrice.objects.get(price=float(amount)).bundle_volume
        else:
            bundle = models.BigTimeBundlePrice.objects.get(price=float(amount)).bundle_volume
        print(bundle)
        new_mtn_transaction = models.BigTimeTransaction.objects.create(
            user=request.user,
            bundle_number=phone_number,
            offer=f"{bundle}MB",
            reference=reference,
        )
        new_mtn_transaction.save()
        user.wallet -= float(amount)
        user.save()
        return JsonResponse({'status': "Your transaction will be completed shortly", 'icon': 'success'})
    return redirect('big_time')


@login_required(login_url='login')
def mtn(request):
    user = models.CustomUser.objects.get(id=request.user.id)
    status = user.status
    form = forms.MTNForm(status)
    db_user_id = request.user.id
    reference = helper.ref_generator()
    user_email = request.user.email

    if request.method == "POST":
        payment_reference = request.POST.get("reference")
        amount_paid = request.POST.get("amount")
        new_payment = models.Payment.objects.create(
            user=request.user,
            reference=payment_reference,
            amount=amount_paid,
            transaction_date=datetime.now(),
            transaction_status="Pending"
        )
        new_payment.save()
        phone_number = request.POST.get("phone")
        offer = request.POST.get("amount")
        if user.status == "User":
            bundle = models.MTNBundlePrice.objects.get(price=float(offer)).bundle_volume
        elif user.status == "Agent":
            bundle = models.AgentMTNBundlePrice.objects.get(price=float(offer)).bundle_volume
        elif user.status == "Super Agent":
            bundle = models.SuperAgentMTNBundlePrice.objects.get(price=float(offer)).bundle_volume
        else:
            bundle = models.MTNBundlePrice.objects.get(price=float(offer)).bundle_volume
        print(phone_number)
        new_mtn_transaction = models.MTNTransaction.objects.create(
            user=request.user,
            bundle_number=phone_number,
            offer=f"{bundle}MB",
            reference=payment_reference,
        )
        new_mtn_transaction.save()
        return JsonResponse({'status': "Your transaction will be completed shortly", 'icon': 'success'})
    user = models.CustomUser.objects.get(id=request.user.id)
    phone_num = user.phone
    mtn_dict = {}

    if user.status == "Agent":
        mtn_offer = models.AgentMTNBundlePrice.objects.all()
    else:
        mtn_offer = models.MTNBundlePrice.objects.all()
    for offer in mtn_offer:
        mtn_dict[str(offer)] = offer.bundle_volume
    context = {'form': form, 'phone_num': phone_num, 'id': db_user_id, 'mtn_dict': json.dumps(mtn_dict),
               "ref": reference, "email": user_email, "wallet": 0 if user.wallet is None else user.wallet}
    return render(request, "layouts/services/mtn.html", context=context)


@login_required(login_url='login')
def afa_registration(request):
    user = models.CustomUser.objects.get(id=request.user.id)
    reference = helper.ref_generator()
    db_user_id = request.user.id
    price = models.AdminInfo.objects.filter().first().afa_price
    user_email = request.user.email
    print(price)
    if request.method == "POST":
        form = forms.AFARegistrationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Registration will be done shortly")
    form = forms.AFARegistrationForm()
    context = {'form': form, 'ref': reference, 'price': price, 'id': db_user_id, "email": user_email,
               "wallet": 0 if user.wallet is None else user.wallet}
    return render(request, "layouts/services/afa.html", context=context)


def afa_registration_wallet(request):
    if request.method == "POST":
        user = models.CustomUser.objects.get(id=request.user.id)
        phone_number = request.POST.get("phone")
        amount = request.POST.get("amount")
        reference = request.POST.get("reference")
        name = request.POST.get("name")
        card_number = request.POST.get("card")
        occupation = request.POST.get("occupation")
        date_of_birth = request.POST.get("birth")
        price = models.AdminInfo.objects.filter().first().afa_price

        if user.wallet is None:
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge.'})
        elif user.wallet <= 0 or user.wallet < float(amount):
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge.'})

        new_registration = models.AFARegistration2.objects.create(
            user=user,
            reference=reference,
            name=name,
            phone_number=phone_number,
            gh_card_number=card_number,
            occupation=occupation,
            date_of_birth=date_of_birth
        )
        new_registration.save()
        user.wallet -= float(price)
        user.save()
        return JsonResponse({'status': "Your transaction will be completed shortly", 'icon': 'success'})
    return redirect('home')


@login_required(login_url='login')
def big_time(request):
    user = models.CustomUser.objects.get(id=request.user.id)
    status = user.status
    form = forms.BigTimeBundleForm(status)
    reference = helper.ref_generator()
    db_user_id = request.user.id
    user_email = request.user.email

    if request.method == "POST":
        payment_reference = request.POST.get("reference")
        amount_paid = request.POST.get("amount")
        new_payment = models.Payment.objects.create(
            user=request.user,
            reference=payment_reference,
            amount=amount_paid,
            transaction_date=datetime.now(),
            transaction_status="Pending"
        )
        new_payment.save()
        phone_number = request.POST.get("phone")
        offer = request.POST.get("amount")
        if user.status == "User":
            bundle = models.BigTimeBundlePrice.objects.get(price=float(offer)).bundle_volume
        elif user.status == "Agent":
            bundle = models.AgentBigTimeBundlePrice.objects.get(price=float(offer)).bundle_volume
        elif user.status == "Super Agent":
            bundle = models.SuperAgentBigTimeBundlePrice.objects.get(price=float(offer)).bundle_volume
        else:
            bundle = models.SuperAgentBigTimeBundlePrice.objects.get(price=float(offer)).bundle_volume
        print(phone_number)
        new_mtn_transaction = models.BigTimeTransaction.objects.create(
            user=request.user,
            bundle_number=phone_number,
            offer=f"{bundle}MB",
            reference=payment_reference,
        )
        new_mtn_transaction.save()
        return JsonResponse({'status': "Your transaction will be completed shortly", 'icon': 'success'})
    user = models.CustomUser.objects.get(id=request.user.id)
    # phone_num = user.phone
    # mtn_dict = {}
    #
    # if user.status == "Agent":
    #     mtn_offer = models.AgentMTNBundlePrice.objects.all()
    # else:
    #     mtn_offer = models.MTNBundlePrice.objects.all()
    # for offer in mtn_offer:
    #     mtn_dict[str(offer)] = offer.bundle_volume
    context = {'form': form,
               "ref": reference, "email": user_email, 'id': db_user_id,
               "wallet": 0 if user.wallet is None else user.wallet}
    return render(request, "layouts/services/big_time.html", context=context)


@login_required(login_url='login')
def history(request):
    user_transactions = models.IShareBundleTransaction.objects.filter(user=request.user).order_by(
        'transaction_date').reverse()
    header = "AirtelTigo Transactions"
    net = "tigo"
    context = {'txns': user_transactions, "header": header, "net": net}
    return render(request, "layouts/history.html", context=context)


@login_required(login_url='login')
def mtn_history(request):
    user_transactions = models.MTNTransaction.objects.filter(user=request.user).order_by('transaction_date').reverse()
    header = "MTN Transactions"
    net = "mtn"
    context = {'txns': user_transactions, "header": header, "net": net}
    return render(request, "layouts/history.html", context=context)


@login_required(login_url='login')
def at_credit_history(request):
    user_transactions = models.ATMinuteTransaction.objects.filter(user=request.user).order_by(
        'transaction_date').reverse()
    header = "AT Minutes Transactions"
    net = "at_min"
    context = {'txns': user_transactions, "header": header, "net": net}
    return render(request, "layouts/history.html", context=context)


@login_required(login_url='login')
def afa_credit_history(request):
    user_transactions = models.AfaCreditTransaction.objects.filter(user=request.user).order_by(
        'transaction_date').reverse()
    header = "Afa Minutes Transaction"
    net = "afa_min"
    context = {'txns': user_transactions, "header": header, "net": net}
    return render(request, "layouts/history.html", context=context)


@login_required(login_url='login')
def big_time_history(request):
    user_transactions = models.BigTimeTransaction.objects.filter(user=request.user).order_by(
        'transaction_date').reverse()
    header = "Big Time Transactions"
    net = "bt"
    context = {'txns': user_transactions, "header": header, "net": net}
    return render(request, "layouts/history.html", context=context)


@login_required(login_url='login')
def afa_history(request):
    user_transactions = models.AFARegistration2.objects.filter(user=request.user).order_by('transaction_date').reverse()
    header = "AFA Registrations"
    net = "afa"
    context = {'txns': user_transactions, "header": header, "net": net}
    return render(request, "layouts/afa_history.html", context=context)


def verify_transaction(request, reference):
    if request.method == "GET":
        response = helper.verify_paystack_transaction(reference)
        data = response.json()
        try:
            status = data["data"]["status"]
            amount = data["data"]["amount"]
            api_reference = data["data"]["reference"]
            date = data["data"]["paid_at"]
            real_amount = float(amount) / 100
            print(status)
            print(real_amount)
            print(api_reference)
            print(reference)
            print(date)
        except:
            status = data["status"]
        return JsonResponse({'status': status})


def change_excel_status(request, status, to_change_to):
    transactions = models.MTNTransaction.objects.filter(
        transaction_status=status) if to_change_to != "Completed" else models.MTNTransaction.objects.filter(
        transaction_status=status).order_by('transaction_date')[:10]
    for transaction in transactions:
        transaction.transaction_status = to_change_to
        transaction.save()
        if to_change_to == "Completed":
            transaction_number = transaction.user.phone
            sms_headers = {
                'Authorization': 'Bearer 1334|wroIm5YnQD6hlZzd8POtLDXxl4vQodCZNorATYGX',
                'Content-Type': 'application/json'
            }

            sms_url = 'https://webapp.usmsgh.com/api/sms/send'
            sms_message = f"Your MTN transaction has been completed. {transaction.bundle_number} has been credited with {transaction.offer}.\nTransaction Reference: {transaction.reference}"

            sms_body = {
                'recipient': f"233{transaction_number}",
                'sender_id': 'GH BAY',
                'message': sms_message
            }
            try:
                response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
                print(response.text)
            except:
                messages.success(request, f"Transaction Completed")
                return redirect('mtn_admin', status=status)
        else:
            messages.success(request, f"Status changed from {status} to {to_change_to}")
            return redirect("mtn_admin", status=status)
    messages.success(request, f"Status changed from {status} to {to_change_to}")
    return redirect("mtn_admin", status=status)


from django.db.models import FloatField
from django.db.models.functions import Cast, Substr, Length


@login_required(login_url='login')
def admin_mtn_history(request, status):
    if request.user.is_staff and request.user.is_superuser:
        if request.method == "POST":
            from io import BytesIO
            from openpyxl import load_workbook
            from django.http import HttpResponse
            import datetime

            # Assuming `uploaded_file` is the Excel file uploaded by the user
            uploaded_file = request.FILES['file'] if 'file' in request.FILES else None
            if not uploaded_file:
                messages.error(request, "No excel file found")
                return redirect('mtn_admin', status=status)

            # Load the uploaded Excel file into memory
            excel_buffer = BytesIO(uploaded_file.read())
            book = load_workbook(excel_buffer)
            sheet = book.active  # Assuming the data is on the active sheet

            # Assuming we have identified the recipient and data column indices
            # Replace these with the actual indices if available
            recipient_col_index = 1  # Example index for "RECIPIENT"
            data_col_index = 2  # Example index for "DATA"

            # Query your Django model
            queryset = models.MTNTransaction.objects.filter(transaction_status="Pending").annotate(
                offer_value=Cast(Substr('offer', 1, Length('offer') - 2), FloatField())
            ).order_by('-offer_value')

            # Determine the starting row for updates, preserving headers and any other pre-existing content
            start_row = 2  # Assuming data starts from row 2

            for record in queryset:
                # Assuming 'bundle_number' and 'offer' fields exist in your model
                recipient_value = f"0{record.bundle_number}"  # Ensure it's a string to preserve formatting
                data_value = record.offer  # Adjust based on actual field type
                cleaned_data_value = float(data_value.replace('MB', ''))
                data_value_gb = round(float(cleaned_data_value) / 1000, 2)

                # Find next available row (avoid overwriting non-empty rows if necessary)
                while sheet.cell(row=start_row, column=recipient_col_index).value is not None:
                    start_row += 1

                # Update cells
                sheet.cell(row=start_row, column=recipient_col_index, value=recipient_value)
                sheet.cell(row=start_row, column=data_col_index, value=data_value_gb)

                # Update the record status, if necessary
                record.transaction_status = "Processing"
                record.save()

            # Save the modified Excel file to the buffer
            excel_buffer.seek(0)  # Reset buffer position
            book.save(excel_buffer)

            # Prepare the response with the modified Excel file
            excel_buffer.seek(0)  # Reset buffer position to read the content
            response = HttpResponse(excel_buffer.getvalue(),
                                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename={}.xlsx'.format(
                datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S"))

            return response

        all_txns = models.MTNTransaction.objects.filter(transaction_status=status).order_by('-transaction_date')[:800]
        context = {'txns': all_txns, 'status': status}
        return render(request, "layouts/services/mtn_admin.html", context=context)
    else:
        messages.error(request, "Access Denied")
        return redirect('mtn_admin', status=status)


@login_required(login_url='login')
def admin_at_history(request):
    if request.user.is_staff and request.user.is_superuser:
        all_txns = models.IShareBundleTransaction.objects.filter().order_by('-transaction_date')[:1000]
        context = {'txns': all_txns}
        return render(request, "layouts/services/at_admin.html", context=context)


@login_required(login_url='login')
def admin_bt_history(request):
    if request.user.is_staff and request.user.is_superuser:
        all_txns = models.BigTimeTransaction.objects.filter().order_by('-transaction_date')[:1000]
        context = {'txns': all_txns}
        return render(request, "layouts/services/bt_admin.html", context=context)


@login_required(login_url='login')
def admin_afa_history(request):
    if request.user.is_staff and request.user.is_superuser:
        all_txns = models.AFARegistration2.objects.filter().order_by('-transaction_date')[:1000]
        context = {'txns': all_txns}
        return render(request, "layouts/services/afa_admin.html", context=context)


@login_required(login_url='login')
def admin_afa_min_history(request):
    if request.user.is_staff and request.user.is_superuser:
        all_txns = models.AfaCreditTransaction.objects.filter().order_by('-transaction_date')[:1000]
        context = {'txns': all_txns}
        return render(request, "layouts/services/afa_min_admin.html", context=context)


@login_required(login_url='login')
def admin_at_min_history(request):
    if request.user.is_staff and request.user.is_superuser:
        all_txns = models.ATMinuteTransaction.objects.filter().order_by('-transaction_date')[:1000]
        context = {'txns': all_txns}
        return render(request, "layouts/services/at_min_admin.html", context=context)


@login_required(login_url='login')
def mark_as_sent(request, pk):
    if request.user.is_staff and request.user.is_superuser:
        txn = models.MTNTransaction.objects.filter(id=pk).first()
        print(txn)
        txn.transaction_status = "Completed"
        txn.save()
        sms_headers = {
            'Authorization': 'Bearer 1334|wroIm5YnQD6hlZzd8POtLDXxl4vQodCZNorATYGX',
            'Content-Type': 'application/json'
        }

        sms_url = 'https://webapp.usmsgh.com/api/sms/send'
        sms_message = f"Your MTN transaction has been completed. {txn.bundle_number} has been credited with {txn.offer}.\nTransaction Reference: {txn.reference}"

        sms_body = {
            'recipient': f"233{txn.user.phone}",
            'sender_id': 'GH BAY',
            'message': sms_message
        }
        try:
            response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
            print(response.text)
        except:
            messages.success(request, f"Transaction Completed")
            return redirect('mtn_admin', status="Pending")
        messages.success(request, f"Transaction Completed")
        return redirect('mtn_admin', status="Pending")


@login_required(login_url='login')
def at_mark_as_sent(request, pk):
    if request.user.is_staff and request.user.is_superuser:
        txn = models.IShareBundleTransaction.objects.filter(id=pk).first()
        print(txn)
        txn.transaction_status = "Completed"
        txn.save()
        sms_headers = {
            'Authorization': 'Bearer 1334|wroIm5YnQD6hlZzd8POtLDXxl4vQodCZNorATYGX',
            'Content-Type': 'application/json'
        }

        sms_url = 'https://webapp.usmsgh.com/api/sms/send'
        sms_message = f"Your AT transaction has been completed. {txn.bundle_number} has been credited with {txn.offer}.\nTransaction Reference: {txn.reference}"

        sms_body = {
            'recipient': f"233{txn.user.phone}",
            'sender_id': 'GH BAY',
            'message': sms_message
        }
        try:
            response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
            print(response.text)
        except:
            messages.success(request, f"Transaction Completed")
            return redirect('at_admin')
        messages.success(request, f"Transaction Completed")
        return redirect('at_admin')


@login_required(login_url='login')
def bt_mark_as_sent(request, pk):
    if request.user.is_staff and request.user.is_superuser:
        txn = models.BigTimeTransaction.objects.filter(id=pk).first()
        print(txn)
        txn.transaction_status = "Completed"
        txn.save()
        sms_headers = {
            'Authorization': 'Bearer 1334|wroIm5YnQD6hlZzd8POtLDXxl4vQodCZNorATYGX',
            'Content-Type': 'application/json'
        }

        sms_url = 'https://webapp.usmsgh.com/api/sms/send'
        sms_message = f"Your AT BIG TIME transaction has been completed. {txn.bundle_number} has been credited with {txn.offer}.\nTransaction Reference: {txn.reference}"

        sms_body = {
            'recipient': f"233{txn.user.phone}",
            'sender_id': 'GH BAY',
            'message': sms_message
        }
        try:
            response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
            print(response.text)
        except:
            messages.success(request, f"Transaction Completed")
            return redirect('bt_admin')
        messages.success(request, f"Transaction Completed")
        return redirect('bt_admin')


@login_required(login_url='login')
def afa_min_mark_as_sent(request, pk):
    if request.user.is_staff and request.user.is_superuser:
        txn = models.AfaCreditTransaction.objects.filter(id=pk).first()
        print(txn)
        txn.transaction_status = "Completed"
        txn.save()
        sms_headers = {
            'Authorization': 'Bearer 1334|wroIm5YnQD6hlZzd8POtLDXxl4vQodCZNorATYGX',
            'Content-Type': 'application/json'
        }

        sms_url = 'https://webapp.usmsgh.com/api/sms/send'
        sms_message = f"Your AFA Minutes transaction has been completed. {txn.bundle_number} has been credited with {txn.offer}.\nTransaction Reference: {txn.reference}"

        sms_body = {
            'recipient': f"233{txn.user.phone}",
            'sender_id': 'GH BAY',
            'message': sms_message
        }
        try:
            response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
            print(response.text)
        except:
            messages.success(request, f"Transaction Completed")
            return redirect('afa_min_admin')
        messages.success(request, f"Transaction Completed")
        return redirect('afa_min_admin')


@login_required(login_url='login')
def at_min_mark_as_sent(request, pk):
    if request.user.is_staff and request.user.is_superuser:
        txn = models.ATMinuteTransaction.objects.filter(id=pk).first()
        print(txn)
        txn.transaction_status = "Completed"
        txn.save()
        sms_headers = {
            'Authorization': 'Bearer 1334|wroIm5YnQD6hlZzd8POtLDXxl4vQodCZNorATYGX',
            'Content-Type': 'application/json'
        }

        sms_url = 'https://webapp.usmsgh.com/api/sms/send'
        sms_message = f"Your AT Minutes transaction has been completed. {txn.bundle_number} has been credited with {txn.offer}.\nTransaction Reference: {txn.reference}"

        sms_body = {
            'recipient': f"233{txn.user.phone}",
            'sender_id': 'GH BAY',
            'message': sms_message
        }
        try:
            response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
            print(response.text)
        except:
            messages.success(request, f"Transaction Completed")
            return redirect('at_min_admin')
        messages.success(request, f"Transaction Completed")
        return redirect('at_min_admin')


@login_required(login_url='login')
def afa_mark_as_sent(request, pk):
    if request.user.is_staff and request.user.is_superuser:
        txn = models.AFARegistration2.objects.filter(id=pk).first()
        print(txn)
        txn.transaction_status = "Completed"
        txn.save()
        sms_headers = {
            'Authorization': 'Bearer 1334|wroIm5YnQD6hlZzd8POtLDXxl4vQodCZNorATYGX',
            'Content-Type': 'application/json'
        }

        sms_url = 'https://webapp.usmsgh.com/api/sms/send'
        sms_message = f"Your AFA Registration has been completed. {txn.phone_number} has been registered.\nTransaction Reference: {txn.reference}"

        sms_body = {
            'recipient': f"233{txn.user.phone}",
            'sender_id': 'GH BAY',
            'message': sms_message
        }
        response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
        print(response.text)
        messages.success(request, f"Transaction Completed")
        return redirect('afa_admin')


@login_required(login_url='login')
def credit_user(request):
    user = models.CustomUser.objects.get(id=request.user.id)
    if request.user.is_superuser:
        form = forms.CreditUserForm()
        if request.method == "POST":
            form = forms.CreditUserForm(request.POST)
            if form.is_valid():
                user = form.cleaned_data["user"]
                amount = form.cleaned_data["amount"]
                print(user)
                print(amount)
                user_needed = models.CustomUser.objects.get(username=user)
                if user_needed.wallet is None:
                    user_needed.wallet = amount
                else:
                    user_needed.wallet += float(amount)
                user_needed.save()
                print(user_needed.username)
                messages.success(request, "Crediting Successful")
                return redirect('credit_user')
        context = {'form': form}
        return render(request, "layouts/services/credit.html", context=context)
    else:
        messages.success(request, "Access Denied")
        return redirect('home')


@login_required(login_url='login')
def topup_info(request):
    if request.method == "POST":
        admin = models.AdminInfo.objects.filter().first().phone_number
        user = models.CustomUser.objects.get(id=request.user.id)
        amount = request.POST.get("amount")
        print(amount)
        reference = helper.top_up_ref_generator()
        new_topup_request = models.TopUpRequest.objects.create(
            user=request.user,
            amount=amount,
            reference=reference,
        )
        new_topup_request.save()

        sms_headers = {
            'Authorization': 'Bearer 1334|wroIm5YnQD6hlZzd8POtLDXxl4vQodCZNorATYGX',
            'Content-Type': 'application/json'
        }

        sms_url = 'https://webapp.usmsgh.com/api/sms/send'
        sms_message = f"A top up request has been placed.\nGHS{amount} for {user}.\nReference: {reference}"

        sms_body = {
            'recipient': f"233{admin}",
            'sender_id': 'GH BAY',
            'message': sms_message
        }
        # response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
        # print(response.text)
        messages.success(request,
                         f"Your Request has been sent successfully. Kindly go on to pay to {admin} and use the reference stated as reference. Reference: {reference}")
        return redirect("request_successful", reference)
    return render(request, "layouts/topup-info.html")


@login_required(login_url='login')
def request_successful(request, reference):
    admin = models.AdminInfo.objects.filter().first()
    context = {
        "name": admin.name,
        "number": f"0{admin.momo_number}",
        "channel": admin.payment_channel,
        "reference": reference
    }
    return render(request, "layouts/services/request_successful.html", context=context)


def topup_list(request):
    if request.user.is_superuser:
        topup_requests = models.TopUpRequest.objects.all().order_by('date').reverse()[:1000]
        context = {
            'requests': topup_requests,
        }
        return render(request, "layouts/services/topup_list.html", context=context)
    else:
        messages.error(request, "Access Denied")
        return redirect('home')


@login_required(login_url='login')
def credit_user_from_list(request, reference):
    if request.user.is_superuser:
        crediting = models.TopUpRequest.objects.filter(reference=reference).first()
        if crediting.status:
            return redirect('topup_list')
        user = crediting.user
        custom_user = models.CustomUser.objects.get(id=user.id)
        amount = crediting.amount
        print(user)
        print(user.phone)
        print(amount)
        custom_user.wallet += amount
        custom_user.save()
        crediting.status = True
        crediting.credited_at = datetime.now()
        crediting.save()
        sms_headers = {
            'Authorization': 'Bearer 1334|wroIm5YnQD6hlZzd8POtLDXxl4vQodCZNorATYGX',
            'Content-Type': 'application/json'
        }

        sms_url = 'https://webapp.usmsgh.com/api/sms/send'
        sms_message = f"Hello,\nYour wallet has been topped up with GHS{amount}.\nReference: {reference}.\nThank you"

        sms_body = {
            'recipient': f"233{custom_user.phone}",
            'sender_id': 'GH BAY',
            'message': sms_message
        }
        try:
            response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
            print(response.text)
        except:
            messages.success(request, f"{user} has been credited with {amount}")
            return redirect('topup_list')
        messages.success(request, f"{user} has been credited with {amount}")
        return redirect('topup_list')


@login_required(login_url='login')
def at_mark_completed(request, reference):
    if request.user.is_superuser:
        txn = models.IShareBundleTransaction.objects.filter(reference=reference).first()
        if txn:
            txn.transaction_status = "Completed"
            txn.save()

            number = txn.user.phone
            bundle = txn.offer

            sms_headers = {
                'Authorization': 'Bearer 1334|wroIm5YnQD6hlZzd8POtLDXxl4vQodCZNorATYGX',
                'Content-Type': 'application/json'
            }

            sms_url = 'https://webapp.usmsgh.com/api/sms/send'
            sms_message = f"Hello,\nYour AT transaction has been completed.\n{bundle} sent to {txn.bundle_number}.\nReference: {reference}.\nThank you for using GH BAY"

            sms_body = {
                'recipient': f"233{number}",
                'sender_id': 'GH BAY',
                'message': sms_message
            }
            try:
                response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
                print(response.text)
            except:
                messages.success(request, f"Transaction Completed")
                return redirect('history')
            messages.success(request, f"Transaction Completed")
            return redirect('history')


@login_required(login_url='login')
def mtn_mark_completed(request, reference):
    if request.user.is_superuser:
        txn = models.MTNTransaction.objects.filter(reference=reference).first()
        if txn:
            txn.transaction_status = "Completed"
            txn.save()

            number = txn.user.phone
            bundle = txn.offer

            sms_headers = {
                'Authorization': 'Bearer 1334|wroIm5YnQD6hlZzd8POtLDXxl4vQodCZNorATYGX',
                'Content-Type': 'application/json'
            }

            sms_url = 'https://webapp.usmsgh.com/api/sms/send'
            sms_message = f"Hello,\nYour MTN transaction has been completed.\n{bundle} sent to {txn.bundle_number}.\nReference: {reference}.\nThank you for using GH BAY"

            sms_body = {
                'recipient': f"233{number}",
                'sender_id': 'GH BAY',
                'message': sms_message
            }
            try:
                response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
                print(response.text)
            except:
                messages.success(request, f"Transaction Completed")
                return redirect('mtn_history')
            messages.success(request, f"Transaction Completed")
            return redirect('mtn_history')


@login_required(login_url='login')
def afa_minutes_mark_completed(request, reference):
    if request.user.is_superuser:
        txn = models.AfaCreditTransaction.objects.filter(reference=reference).first()
        if txn:
            txn.transaction_status = "Completed"
            txn.save()

            number = txn.user.phone
            bundle = txn.offer

            sms_headers = {
                'Authorization': 'Bearer 1334|wroIm5YnQD6hlZzd8POtLDXxl4vQodCZNorATYGX',
                'Content-Type': 'application/json'
            }

            sms_url = 'https://webapp.usmsgh.com/api/sms/send'
            sms_message = f"Hello,\nYour Afa Minutes transaction has been completed.\n{bundle} minutes sent to {txn.bundle_number}.\nReference: {reference}.\nThank you for using GH BAY"

            sms_body = {
                'recipient': f"233{number}",
                'sender_id': 'GH BAY',
                'message': sms_message
            }
            try:
                response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
                print(response.text)
            except:
                messages.success(request, f"Transaction Completed")
                return redirect('afa-credit-history')
            messages.success(request, f"Transaction Completed")
            return redirect('afa-credit-history')


@login_required(login_url='login')
def at_minutes_mark_completed(request, reference):
    if request.user.is_superuser:
        txn = models.ATMinuteTransaction.objects.filter(reference=reference).first()
        if txn:
            txn.transaction_status = "Completed"
            txn.save()

            number = txn.user.phone
            bundle = txn.offer

            sms_headers = {
                'Authorization': 'Bearer 1334|wroIm5YnQD6hlZzd8POtLDXxl4vQodCZNorATYGX',
                'Content-Type': 'application/json'
            }

            sms_url = 'https://webapp.usmsgh.com/api/sms/send'
            sms_message = f"Hello,\nYour AT Minutes transaction has been completed.\n{bundle} minutes sent to {txn.bundle_number}.\nReference: {reference}.\nThank you for using GH BAY"

            sms_body = {
                'recipient': f"233{number}",
                'sender_id': 'GH BAY',
                'message': sms_message
            }
            try:
                response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
                print(response.text)
            except:
                messages.success(request, f"Transaction Completed")
                return redirect('afa-credit-history')
            messages.success(request, f"Transaction Completed")
            return redirect('afa-credit-history')


# def import_users(request):
#     if request.method == 'POST':
#         form = UploadFileForm(request.POST, request.FILES)
#         if form.is_valid():
#             file = request.FILES['file']
#             if not file.name.endswith('.xlsx'):
#                 messages.error(request, 'Please upload a valid Excel file.')
#                 return render(request, 'import_users.html', {'form': form})
#
#             dataset = Dataset()
#             imported_data = dataset.load(file.read(), format='xlsx')
#             resource = CustomUserResource()
#             result = resource.import_data(imported_data, dry_run=True)  # Perform a dry run to validate data
#
#             if result.has_errors():
#                 messages.error(request, 'There were errors importing the data. Please check the file and try again.')
#             else:
#                 resource.import_data(imported_data, dry_run=False)  # Import the data into the CustomUser model
#                 messages.success(request, 'User data imported successfully.')
#     else:
#         form = UploadFileForm()
#     return render(request, 'layouts/import_users.html', {'form': form})


def populate_custom_users_from_excel(request):
    # Read the Excel file using pandas
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['file']

            # Process the uploaded Excel file
            df = pd.read_excel(excel_file)
            counter = 0
            # Iterate through rows to create CustomUser instances
            for index, row in df.iterrows():
                print(counter)
                # Create a CustomUser instance for each row
                custom_user = CustomUser.objects.create(
                    first_name=row['first_name'],
                    last_name=row['last_name'],
                    username=str(row['username']),
                    email=row['email'],
                    phone=row['phone'],
                    wallet=float(row['wallet']),
                    status=str(row['status']),
                    password1=row['password1'],
                    password2=row['password2'],
                    is_superuser=row['is_superuser'],
                    is_staff=row['is_staff'],
                    is_active=row['is_active'],
                    password=row['password']
                )

                custom_user.save()

                # group_names = row['groups'].split(',')  # Assuming groups are comma-separated
                # groups = Group.objects.filter(name__in=group_names)
                # custom_user.groups.set(groups)
                #
                # if row['user_permissions']:
                #     permission_ids = [int(pid) for pid in row['user_permissions'].split(',')]
                #     permissions = Permission.objects.filter(id__in=permission_ids)
                #     custom_user.user_permissions.set(permissions)
                print("killed")
                counter = counter + 1
            messages.success(request, 'All done')
    else:
        form = UploadFileForm()
    return render(request, 'layouts/import_users.html', {'form': form})


@csrf_exempt
def export_unknown_transactions(request):
    existing_excel_path = 'wallet_api_app/ALL PACKAGES LATEST.xlsx'  # Update with your file path

    # Load the existing Excel file using openpyxl.Workbook
    book = load_workbook(existing_excel_path)

    # Get the active sheet
    sheet_name = 'Sheet1'
    sheet = book[sheet_name] if sheet_name in book.sheetnames else book.active

    # Clear existing data from the sheet (excluding headers)
    for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, max_row=sheet.max_row):
        for cell in row:
            cell.value = None

    # Query your Django model for the first 200 records with batch_id 'Unknown' and ordered by status and date
    queryset = models.MTNTransaction.objects.filter(batch_id='Unknown', status="Undelivered")[:50]

    # Process transactions with batch_id 'Unknown'
    counter = 0

    for record in queryset:
        print(counter)

        # Extract required fields from your Django model
        bundle_volume_mb = record.bundle_volume  # Assuming a default of 0 if datavolume is missing
        number = f"0{record.number}"  # Convert to string to keep leading zeros

        # Convert datavolume from MB to GB
        bundle_volume_gb = round(float(bundle_volume_mb) / 1000)

        # Find the row index where you want to populate the data (adjust as needed)
        target_row = 2 + counter  # Assuming the data starts from row 2

        # Populate the specific cells with the new data
        sheet.cell(row=target_row, column=1, value=str(number))  # Keep leading zeros
        sheet.cell(row=target_row, column=2, value=float(bundle_volume_gb))  # Convert to float

        # Update 'batch_id' to 'processing' in your Django model
        record.batch_id = 'accepted'
        record.status = 'Processing'
        record.save()

        counter += 1

    print(f"Total transactions to export: {counter}")

    # Save changes to the existing Excel file
    book.save(existing_excel_path)

    # You can continue with the response as needed
    excel_buffer = BytesIO()

    # Save the workbook to the buffer
    book.save(excel_buffer)

    # Create a response with the Excel file
    response = HttpResponse(excel_buffer.getvalue(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={datetime.datetime.now()}.xlsx'

    return response


def delete_custom_users(request):
    CustomUser.objects.all().delete()
    return HttpResponseRedirect('Done')


def send_change_sms(request):
    sms_headers = {
        'Authorization': 'Bearer 1334|wroIm5YnQD6hlZzd8POtLDXxl4vQodCZNorATYGX',
        'Content-Type': 'application/json'
    }

    sms_url = 'https://webapp.usmsgh.com/api/sms/send'

    all_users = CustomUser.objects.all()
    counter = 0

    for user in all_users:
        sleep(1)
        sms_message = f"Hello {user.username},\nGH Bay has changed its website to https://www.ghbays.com\nYou do not need to create a new account for this. Just log in and start transacting\nAll wallet balance are intact."

        sms_body = {
            'recipient': f"233{user.phone}",
            'sender_id': 'GH BAY',
            'message': sms_message
        }

        response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
        print(response.text)
        counter = counter + 1
        print(counter)
        print("killed")
    messages.success(request, "ALL DONE")
    return redirect('home')


@login_required(login_url='login')
def pay_with_wallet_minutes(request):
    if request.method == "POST":
        admin = models.AdminInfo.objects.filter().first().phone_number
        user = models.CustomUser.objects.get(id=request.user.id)
        phone_number = request.POST.get("phone")
        amount = request.POST.get("amount")
        reference = request.POST.get("reference")
        if user.wallet is None:
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge. Admin Contact Info: 0{admin}'})
        elif user.wallet <= 0 or user.wallet < float(amount):
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge. Admin Contact Info: 0{admin}'})
        print(phone_number)
        print(amount)
        print(reference)
        if user.status == "User":
            minutes = models.ATCreditPrice.objects.get(price=float(amount)).minutes
        else:
            minutes = models.ATCreditPrice.objects.get(price=float(amount)).minutes

        new_transaction = models.ATMinuteTransaction.objects.create(
            user=request.user,
            bundle_number=phone_number,
            offer=f"{minutes} Minutes",
            reference=reference,
            transaction_status="Pending"
        )
        new_transaction.save()
        user.wallet -= float(amount)
        user.save()

        return JsonResponse({'status': 'Transaction Received Successfully', 'icon': 'success'})
    return redirect('at_minutes')


@login_required(login_url='login')
def afa_credit_pay_with_wallet(request):
    if request.method == "POST":
        admin = models.AdminInfo.objects.filter().first().phone_number
        user = models.CustomUser.objects.get(id=request.user.id)
        phone_number = request.POST.get("phone")
        amount = request.POST.get("amount")
        reference = request.POST.get("reference")
        if user.wallet is None:
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge. Admin Contact Info: 0{admin}'})
        elif user.wallet <= 0 or user.wallet < float(amount):
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge. Admin Contact Info: 0{admin}'})
        print(phone_number)
        print(amount)
        print(reference)
        if user.status == "User":
            minutes = models.AfaCreditPrice.objects.get(price=float(amount)).minutes
        else:
            minutes = models.AfaCreditPrice.objects.get(price=float(amount)).minutes

        new_transaction = models.AfaCreditTransaction.objects.create(
            user=request.user,
            bundle_number=phone_number,
            offer=f"{minutes} Minutes",
            reference=reference,
            transaction_status="Pending"
        )
        new_transaction.save()
        user.wallet -= float(amount)
        user.save()

        return JsonResponse({'status': 'Transaction Received Successfully', 'icon': 'success'})
    return redirect('afa_credit')


@login_required(login_url='login')
def airtel_tigo_minutes(request):
    user = models.CustomUser.objects.get(id=request.user.id)
    status = user.status
    form = forms.ATCreditForm()
    reference = helper.ref_generator()
    db_user_id = request.user.id
    user_email = request.user.email
    if request.method == "POST":
        form = forms.ATCreditForm(data=request.POST)
        payment_reference = request.POST.get("reference")
        amount_paid = request.POST.get("amount")
        new_payment = models.Payment.objects.create(
            user=request.user,
            reference=payment_reference,
            amount=amount_paid,
            transaction_date=datetime.now(),
            transaction_status="Completed"
        )
        new_payment.save()
        print("payment saved")
        print("form valid")
        phone_number = request.POST.get("phone")
        offer = request.POST.get("amount")
        print(offer)
        if user.status == "User":
            bundle = models.ATCreditPrice.objects.get(price=float(offer)).minutes
        else:
            bundle = models.ATCreditPrice.objects.get(price=float(offer)).minutes
        new_transaction = models.ATMinuteTransaction.objects.create(
            user=request.user,
            bundle_number=phone_number,
            offer=f"{bundle} Minutes",
            reference=payment_reference,
            transaction_status="Pending"
        )
        print("created")
        new_transaction.save()

        print("===========================")
        print(phone_number)
        print(bundle)
        print("--------------------")
        # send_bundle_response = helper.send_bundle(request.user, phone_number, bundle, payment_reference)
        # print("********************")
        # data = send_bundle_response.json()
        #
        # print(data)
        return JsonResponse({'status': 'Transaction Completed Successfully', 'icon': 'success'})
    user = models.CustomUser.objects.get(id=request.user.id)
    context = {"form": form, "ref": reference, 'id': db_user_id, "email": user_email,
               "wallet": 0 if user.wallet is None else user.wallet}
    return render(request, "layouts/services/at_credit.html", context=context)


@login_required(login_url='login')
def afa_credit(request):
    user = models.CustomUser.objects.get(id=request.user.id)
    status = user.status
    form = forms.AfaCreditForm()
    reference = helper.ref_generator()
    db_user_id = request.user.id
    user_email = request.user.email
    if request.method == "POST":
        form = forms.AfaCreditForm(data=request.POST)
        payment_reference = request.POST.get("reference")
        amount_paid = request.POST.get("amount")
        new_payment = models.Payment.objects.create(
            user=request.user,
            reference=payment_reference,
            amount=amount_paid,
            transaction_date=datetime.now(),
            transaction_status="Completed"
        )
        new_payment.save()
        print("payment saved")
        print("form valid")
        phone_number = request.POST.get("phone")
        offer = request.POST.get("amount")
        print(offer)
        if user.status == "User":
            bundle = models.AfaCreditPrice.objects.get(price=float(offer)).minutes
        else:
            bundle = models.AfaCreditPrice.objects.get(price=float(offer)).minutes
        new_transaction = models.AfaCreditTransaction.objects.create(
            user=request.user,
            bundle_number=phone_number,
            offer=f"{bundle} Minutes",
            reference=payment_reference,
            transaction_status="Pending"
        )
        print("created")
        new_transaction.save()

        print("===========================")
        print(phone_number)
        print(bundle)
        print("--------------------")
        # send_bundle_response = helper.send_bundle(request.user, phone_number, bundle, payment_reference)
        # print("********************")
        # data = send_bundle_response.json()
        #
        # print(data)
        return JsonResponse({'status': 'Transaction Received Successfully', 'icon': 'success'})
    user = models.CustomUser.objects.get(id=request.user.id)
    context = {"form": form, "ref": reference, 'id': db_user_id, "email": user_email,
               "wallet": 0 if user.wallet is None else user.wallet}
    return render(request, "layouts/services/afa_credit.html", context=context)


@csrf_exempt
def paystack_webhook(request):
    if request.method == "POST":
        paystack_secret_key = config("PAYSTACK_SECRET_KEY")
        # print(paystack_secret_key)
        payload = json.loads(request.body)

        paystack_signature = request.headers.get("X-Paystack-Signature")

        if not paystack_secret_key or not paystack_signature:
            return HttpResponse(status=400)

        computed_signature = hmac.new(
            paystack_secret_key.encode('utf-8'),
            request.body,
            hashlib.sha512
        ).hexdigest()

        if computed_signature == paystack_signature:
            print("yes")
            print(payload.get('data'))
            r_data = payload.get('data')
            print(r_data.get('metadata'))
            print(payload.get('event'))
            if payload.get('event') == 'charge.success':
                metadata = r_data.get('metadata')
                receiver = metadata.get('receiver')
                db_id = metadata.get('db_id')
                print(db_id)
                # offer = metadata.get('offer')
                user = models.CustomUser.objects.get(id=int(db_id))
                print(user)
                channel = metadata.get('channel')
                real_amount = metadata.get('real_amount')
                print(real_amount)
                paid_amount = r_data.get('amount')
                reference = r_data.get('reference')

                paid_amount = r_data.get('amount')
                reference = r_data.get('reference')

                slashed_amount = float(paid_amount) / 100
                reference = r_data.get('reference')

                if channel != "afa":
                    rounded_real_amount = round(float(real_amount))
                    rounded_paid_amount = round(float(slashed_amount))

                    print(f"reeeeeeeaaaaaaaaal amount: {rounded_real_amount}")
                    print(f"paaaaaaaaaaaaaiiddd amount: {rounded_paid_amount}")

                    is_within_range = (rounded_real_amount - 3) <= rounded_paid_amount <= (rounded_real_amount + 3)

                    if not is_within_range:
                        sms_headers = {
                            'Authorization': 'Bearer 1334|wroIm5YnQD6hlZzd8POtLDXxl4vQodCZNorATYGX',
                            'Content-Type': 'application/json'
                        }

                        sms_url = 'https://webapp.usmsgh.com/api/sms/send'
                        sms_message = f"Malicious attempt on webhook. Real amount: {rounded_real_amount} | Paid amount: {rounded_paid_amount}. Referrer: {reference}"

                        sms_body = {
                            'recipient': "233242442147",
                            'sender_id': 'GH BAY',
                            'message': sms_message
                        }
                        try:
                            response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
                            print(response.text)
                        except:
                            pass

                        print("not within range")
                        return HttpResponse(200)

                if channel == "ishare":
                    if user.status == "User":
                        bundle = models.IshareBundlePrice.objects.get(price=float(real_amount)).bundle_volume
                    elif user.status == "Agent":
                        bundle = models.AgentIshareBundlePrice.objects.get(price=float(real_amount)).bundle_volume
                    elif user.status == "Super Agent":
                        bundle = models.SuperAgentIshareBundlePrice.objects.get(price=float(real_amount)).bundle_volume
                    else:
                        bundle = models.IshareBundlePrice.objects.get(price=float(real_amount)).bundle_volume

                    if models.IShareBundleTransaction.objects.filter(reference=reference, offer=f"{bundle}MB",
                                                                     transaction_status="Completed").exists():
                        return HttpResponse(status=200)


                    else:
                        send_bundle_response = helper.send_bundle(receiver, bundle, reference)
                        try:
                            data = send_bundle_response.json()
                            print(data)
                        except:
                            return HttpResponse(status=500)

                        sms_headers = {
                            'Authorization': 'Bearer 1334|wroIm5YnQD6hlZzd8POtLDXxl4vQodCZNorATYGX',
                            'Content-Type': 'application/json'
                        }

                        sms_url = 'https://webapp.usmsgh.com/api/sms/send'
                        if send_bundle_response.status_code == 200:
                            if data["status"] == "Success":
                                new_transaction = models.IShareBundleTransaction.objects.create(
                                    user=user,
                                    bundle_number=receiver,
                                    offer=f"{bundle}MB",
                                    reference=reference,
                                    transaction_status="Completed"
                                )
                                new_transaction.save()
                                receiver_message = f"Your bundle purchase has been completed successfully. {bundle}MB has been credited to you by {user.phone}.\nReference: {reference}\n"
                                sms_message = f"Hello @{request.user.username}. Your bundle purchase has been completed successfully. {bundle}MB has been credited to {receiver}.\nReference: {reference}\nCurrent Wallet Balance: {user.wallet}\nThank you for using GH BAY."

                                num_without_0 = receiver[1:]
                                print(num_without_0)
                                receiver_body = {
                                    'recipient': f"233{num_without_0}",
                                    'sender_id': 'GH BAY',
                                    'message': receiver_message
                                }

                                response = requests.request('POST', url=sms_url, params=receiver_body,
                                                            headers=sms_headers)
                                print(response.text)

                                sms_body = {
                                    'recipient': f"233{request.user.phone}",
                                    'sender_id': 'GH BAY',
                                    'message': sms_message
                                }

                                response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)

                                print(response.text)
                                return HttpResponse(status=200)
                            else:
                                new_transaction = models.IShareBundleTransaction.objects.create(
                                    user=user,
                                    bundle_number=receiver,
                                    offer=f"{bundle}MB",
                                    reference=reference,
                                    transaction_status="Pending"
                                )
                                new_transaction.save()
                                return HttpResponse(status=500)
                        else:
                            new_transaction = models.IShareBundleTransaction.objects.create(
                                user=user,
                                bundle_number=receiver,
                                offer=f"{bundle}MB",
                                reference=reference,
                                transaction_status="Pending"
                            )
                            new_transaction.save()
                            return HttpResponse(status=500)
                elif channel == "mtn":
                    new_payment = models.Payment.objects.create(
                        user=user,
                        reference=reference,
                        amount=paid_amount,
                        transaction_date=datetime.now(),
                        transaction_status="Pending"
                    )
                    new_payment.save()

                    if user.status == "User":
                        bundle = models.MTNBundlePrice.objects.get(price=float(real_amount)).bundle_volume
                    elif user.status == "Agent":
                        bundle = models.AgentMTNBundlePrice.objects.get(price=float(real_amount)).bundle_volume
                    elif user.status == "Super Agent":
                        bundle = models.SuperAgentMTNBundlePrice.objects.get(price=float(real_amount)).bundle_volume
                    else:
                        bundle = models.MTNBundlePrice.objects.get(price=float(real_amount)).bundle_volume

                    print(receiver)

                    new_mtn_transaction = models.MTNTransaction.objects.create(
                        user=user,
                        bundle_number=receiver,
                        offer=f"{bundle}MB",
                        reference=reference,
                    )
                    new_mtn_transaction.save()
                    return HttpResponse(status=200)
                elif channel == "at_min":
                    new_payment = models.Payment.objects.create(
                        user=user,
                        reference=reference,
                        amount=paid_amount,
                        transaction_date=datetime.now(),
                        transaction_status="Pending"
                    )
                    new_payment.save()

                    if user.status == "User":
                        minutes = models.ATCreditPrice.objects.get(price=float(real_amount)).minutes
                    else:
                        minutes = models.ATCreditPrice.objects.get(price=float(real_amount)).minutes

                    print(receiver)

                    new_mtn_transaction = models.ATMinuteTransaction.objects.create(
                        user=user,
                        bundle_number=receiver,
                        offer=f"{minutes} Minutes",
                        reference=reference,
                    )
                    new_mtn_transaction.save()
                    return HttpResponse(status=200)
                elif channel == "afa_credit":
                    new_payment = models.Payment.objects.create(
                        user=user,
                        reference=reference,
                        amount=paid_amount,
                        transaction_date=datetime.now(),
                        transaction_status="Pending"
                    )
                    new_payment.save()

                    if user.status == "User":
                        minutes = models.AfaCreditPrice.objects.get(price=float(real_amount)).minutes
                    else:
                        minutes = models.AfaCreditPrice.objects.get(price=float(real_amount)).minutes

                    print(receiver)

                    new_mtn_transaction = models.AfaCreditTransaction.objects.create(
                        user=user,
                        bundle_number=receiver,
                        offer=f"{minutes} Minutes",
                        reference=reference,
                    )
                    new_mtn_transaction.save()
                    return HttpResponse(status=200)
                elif channel == "big-time":
                    new_payment = models.Payment.objects.create(
                        user=user,
                        reference=reference,
                        amount=paid_amount,
                        transaction_date=datetime.now(),
                        transaction_status="Pending"
                    )
                    new_payment.save()

                    if user.status == "User":
                        bundle = models.BigTimeBundlePrice.objects.get(price=float(real_amount)).bundle_volume
                    elif user.status == "Agent":
                        bundle = models.AgentBigTimeBundlePrice.objects.get(price=float(real_amount)).bundle_volume
                    elif user.status == "Super Agent":
                        bundle = models.SuperAgentBigTimeBundlePrice.objects.get(price=float(real_amount)).bundle_volume
                    else:
                        bundle = models.BigTimeBundlePrice.objects.get(price=float(real_amount)).bundle_volume

                    print(receiver)

                    new_transaction = models.BigTimeTransaction.objects.create(
                        user=user,
                        bundle_number=receiver,
                        offer=f"{bundle}MB",
                        reference=reference,
                    )
                    new_transaction.save()
                    return HttpResponse(status=200)
                elif channel == "afa":
                    phone_number = metadata.get('phone_number')
                    gh_card_number = metadata.get('card_number')
                    name = metadata.get('name')
                    occupation = metadata.get('occupation')
                    date_of_birth = metadata.get('dob')

                    new_payment = models.Payment.objects.create(
                        user=user,
                        reference=reference,
                        amount=paid_amount,
                        transaction_date=datetime.now(),
                        transaction_status="Pending"
                    )
                    new_payment.save()

                    new_afa_txn = models.AFARegistration2.objects.create(
                        user=user,
                        reference=reference,
                        name=name,
                        gh_card_number=gh_card_number,
                        phone_number=phone_number,
                        occupation=occupation,
                        date_of_birth=date_of_birth
                    )
                    new_afa_txn.save()
                    return HttpResponse(status=200)
                elif channel == "commerce":
                    phone_number = metadata.get('phone_number')
                    region = metadata.get('region')
                    name = metadata.get('name')
                    city = metadata.get('city')
                    message = metadata.get('message')
                    address = metadata.get('address')
                    order_mail = metadata.get('order_mail')

                    print(phone_number, region, name, city, message, address, order_mail)

                    new_order_items = models.Cart.objects.filter(user=user)
                    cart = models.Cart.objects.filter(user=user)
                    cart_total_price = 0
                    for item in cart:
                        cart_total_price += item.product.selling_price * item.product_qty
                    print(cart_total_price)
                    print(user.wallet)
                    if models.Order.objects.filter(tracking_number=reference, message=message,
                                                   payment_id=reference).exists():
                        return HttpResponse(status=200)
                    order_form = models.Order.objects.create(
                        user=user,
                        full_name=name,
                        email=order_mail,
                        phone=phone_number,
                        address=address,
                        city=city,
                        region=region,
                        total_price=cart_total_price,
                        payment_mode="Paystack",
                        payment_id=reference,
                        message=message,
                        tracking_number=reference
                    )
                    order_form.save()

                    for item in new_order_items:
                        models.OrderItem.objects.create(
                            order=order_form,
                            product=item.product,
                            tracking_number=order_form.tracking_number,
                            price=item.product.selling_price,
                            quantity=item.product_qty
                        )
                        order_product = models.Product.objects.filter(id=item.product_id).first()
                        order_product.quantity -= item.product_qty
                        order_product.save()

                    models.Cart.objects.filter(user=user).delete()

                    sms_headers = {
                        'Authorization': 'Bearer 1334|wroIm5YnQD6hlZzd8POtLDXxl4vQodCZNorATYGX',
                        'Content-Type': 'application/json'
                    }

                    sms_url = 'https://webapp.usmsgh.com/api/sms/send'
                    sms_message = f"Order Placed Successfully\nYour order with order number {order_form.tracking_number} has been received and is being processed.\nYou will receive a message when your order is Out for Delivery.\nThank you for shopping with GH BAY"

                    sms_body = {
                        'recipient': f"233{order_form.phone}",
                        'sender_id': 'GH BAY',
                        'message': sms_message
                    }
                    try:
                        response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
                        print(response.text)
                    except:
                        print("Could not send sms message")
                    return HttpResponse(status=200)
                elif channel == "voda":
                    new_payment = models.Payment.objects.create(
                        user=user,
                        reference=reference,
                        amount=paid_amount,
                        transaction_date=datetime.now(),
                        transaction_status="Completed"
                    )
                    new_payment.save()

                    if user.status == "User":
                        bundle = models.VodaBundlePrice.objects.get(price=float(real_amount)).bundle_volume
                    elif user.status == "Agent":
                        bundle = models.AgentVodaBundlePrice.objects.get(price=float(real_amount)).bundle_volume
                    elif user.status == "Super Agent":
                        bundle = models.SuperAgentVodaBundlePrice.objects.get(price=float(real_amount)).bundle_volume
                    else:
                        bundle = models.VodaBundlePrice.objects.get(price=float(real_amount)).bundle_volume

                    new_voda_transaction = models.VodafoneTransaction.objects.create(
                        user=user,
                        bundle_number=receiver,
                        offer=f"{bundle}MB",
                        reference=reference,
                    )
                    new_voda_transaction.save()

                    return HttpResponse(status=200)
                else:
                    return HttpResponse(status=200)
            else:
                return HttpResponse(status=200)
        else:
            return HttpResponse(status=401)


@login_required(login_url='login')
def voda(request):
    user = models.CustomUser.objects.get(id=request.user.id)
    status = user.status
    form = forms.VodaBundleForm(status)
    reference = helper.ref_generator()
    user_email = request.user.email
    db_user_id = request.user.id

    # if request.method == "POST":
    # payment_reference = request.POST.get("reference")
    # amount_paid = request.POST.get("amount")
    # new_payment = models.Payment.objects.create(
    #     user=request.user,
    #     reference=payment_reference,
    #     amount=amount_paid,
    #     transaction_date=datetime.now(),
    #     transaction_status="Pending"
    # )
    # new_payment.save()
    # phone_number = request.POST.get("phone")
    # offer = request.POST.get("amount")
    # bundle = models.VodaBundlePrice.objects.get(
    #     price=float(offer)).bundle_volume if user.status == "User" else models.AgentVodaBundlePrice.objects.get(
    #     price=float(offer)).bundle_volume
    #
    # print(phone_number)
    # new_mtn_transaction = models.VodafoneTransaction.objects.create(
    #     user=request.user,
    #     bundle_number=phone_number,
    #     offer=f"{bundle}MB",
    #     reference=payment_reference,
    # )
    # new_mtn_transaction.save()
    # return JsonResponse({'status': "Your transaction will be completed shortly", 'icon': 'success'})
    user = models.CustomUser.objects.get(id=request.user.id)
    # phone_num = user.phone
    # mtn_dict = {}
    #
    # if user.status == "Agent":
    #     mtn_offer = models.AgentMTNBundlePrice.objects.all()
    # else:
    #     mtn_offer = models.MTNBundlePrice.objects.all()
    # for offer in mtn_offer:
    #     mtn_dict[str(offer)] = offer.bundle_volume
    context = {'form': form,
               "ref": reference, "email": user_email, "wallet": 0 if user.wallet is None else user.wallet,
               'id': db_user_id}
    return render(request, "layouts/services/voda.html", context=context)


@login_required(login_url='login')
def voda_pay_with_wallet(request):
    if request.method == "POST":
        user = models.CustomUser.objects.get(id=request.user.id)
        phone_number = request.POST.get("phone")
        amount = request.POST.get("amount")
        reference = request.POST.get("reference")
        print(phone_number)
        print(amount)
        print(reference)
        if user.wallet is None:
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge.'})
        elif user.wallet <= 0 or user.wallet < float(amount):
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge.'})

        if user.status == "User":
            bundle = models.VodaBundlePrice.objects.get(price=float(amount)).bundle_volume
        elif user.status == "Agent":
            bundle = models.AgentVodaBundlePrice.objects.get(price=float(amount)).bundle_volume
        elif user.status == "Super Agent":
            bundle = models.SuperAgentVodaBundlePrice.objects.get(price=float(amount)).bundle_volume
        else:
            bundle = models.VodaBundlePrice.objects.get(price=float(amount)).bundle_volume

        print(bundle)
        new_mtn_transaction = models.VodafoneTransaction.objects.create(
            user=request.user,
            bundle_number=phone_number,
            offer=f"{bundle}MB",
            reference=reference,
        )
        new_mtn_transaction.save()
        user.wallet -= float(amount)
        user.save()
        return JsonResponse({'status': "Your transaction will be completed shortly", 'icon': 'success'})
    return redirect('voda')


@login_required(login_url='login')
def voda_history(request):
    user_transactions = models.VodafoneTransaction.objects.filter(user=request.user).order_by(
        'transaction_date').reverse()
    header = "Vodafone Transactions"
    net = "voda"
    context = {'txns': user_transactions, "header": header, "net": net}
    return render(request, "layouts/history.html", context=context)


@login_required(login_url='login')
def admin_voda_history(request):
    if request.user.is_staff and request.user.is_superuser:
        all_txns = models.VodafoneTransaction.objects.filter().order_by('-transaction_date')[:1000]
        context = {'txns': all_txns}
        return render(request, "layouts/services/voda_admin.html", context=context)


@login_required(login_url='login')
def voda_mark_as_sent(request, pk):
    if request.user.is_staff and request.user.is_superuser:
        txn = models.VodafoneTransaction.objects.filter(id=pk).first()
        print(txn)
        txn.transaction_status = "Completed"
        txn.save()
        sms_headers = {
            'Authorization': 'Bearer 1334|wroIm5YnQD6hlZzd8POtLDXxl4vQodCZNorATYGX',
            'Content-Type': 'application/json'
        }

        sms_url = 'https://webapp.usmsgh.com/api/sms/send'
        sms_message = f"Your Vodafone transaction has been completed. {txn.bundle_number} has been credited with {txn.offer}.\nTransaction Reference: {txn.reference}"

        receiver_body = {
            'recipient': f"233{txn.user.phone}",
            'sender_id': 'GH BAY',
            'message': sms_message
        }

        response = requests.request('POST', url=sms_url, params=receiver_body,
                                    headers=sms_headers)
        print(response.text)
        # response1 = requests.get(
        #     f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=Ojd0Uk5BVmE3SUpna2lRS3o=&to=0{txn.user.phone}&from=Noble Data&sms={sms_message}")
        # print(response1.text)
        messages.success(request, f"Transaction Completed")
        return redirect('voda_admin')


# ======================================================================================================================================
# ======================================================================================================================================
# ======================================================================================================================================
# ======================================================================================================================================


def track_shipment(request):
    if request.method == "POST":
        tracking_number = request.POST.get("tracking_number")
        print(tracking_number)

        return redirect('track_order', tracking_number=tracking_number)
    return render(request, "layouts/tracking_goods/tracking_home.html")


from django.shortcuts import render
from .models import Tracking, ShippingOrder, Package


def track_order(request, tracking_number):
    try:
        order = models.ShippingOrder.objects.get(order_number=tracking_number)
        packages = order.packages.all()
        total = sum(package.price for package in packages)
    except Tracking.DoesNotExist:
        order, packages, total = None, None, 0

    return render(request, 'layouts/track_order.html', {
        'order': order,
        'packages': packages,
        'total': total,
        'tracking_number': tracking_number
    })


import random
import string


def generate_tracking_number():
    letters = string.ascii_uppercase
    digits = string.digits
    part1 = ''.join(random.choices(letters, k=3))
    part2 = ''.join(random.choices(digits, k=4))
    part3 = ''.join(random.choices(letters, k=2))
    return f"{part1}-{part2}-{part3}"


def create_order(request):
    if request.method == 'POST':
        order_form = OrderForm(request.POST)
        package_forms = [PackageForm(request.POST, prefix=str(i)) for i in range(50)]

        if order_form.is_valid():
            order = order_form.save()
            for package_form in package_forms:
                if package_form.is_valid() and any(package_form.cleaned_data.values()):
                    package = package_form.save(commit=False)
                    package.order = order
                    package.save()
                else:
                    print("package forms")
                    print(package_form.errors)
            tracking = Tracking.objects.create(
                order=order,
                tracking_number=generate_tracking_number()
            )
            tracking.save()
            messages.success(request, f"Order Created. Tracking number is {order.order_number}")

            sms_url = 'https://webapp.usmsgh.com/api/sms/send'
            sms_message = f"Hello, {order.owner_name}, Your goods have been received in our China warehouse. Track your shipment with the link below.\nhttps://www.ghbays.com/trackshipment/{order.order_number}"

            receiver_body = {
                'recipient': f"233{order.phone_number[1:]}",
                'sender_id': 'GH BAY',
                'message': sms_message
            }

            sms_headers = {
                'Authorization': 'Bearer 1334|wroIm5YnQD6hlZzd8POtLDXxl4vQodCZNorATYGX',
                'Content-Type': 'application/json'
            }

            response = requests.request('POST', url=sms_url, params=receiver_body,
                                        headers=sms_headers)
            print(response.text)
            return redirect('create_order')
        else:
            print(order_form.errors)
    else:
        order_form = OrderForm()
        package_forms = [PackageForm(prefix=str(i)) for i in range(50)]

    return render(request, 'layouts/create_order.html', {
        'order_form': order_form,
        'package_forms': package_forms,
    })


def admin_order_list(request):
    if request.method == 'POST':
        order_id = request.POST.get('order_id')
        status = request.POST.get('status')
        order = ShippingOrder.objects.get(id=order_id)
        tracking = models.Tracking.objects.get(order=order)

        if status == 'Loaded':
            order.loaded_date = timezone.now()
        elif status == 'Received':
            order.received_date = timezone.now()
        elif status == 'Out for Delivery':
            order.estimated_date_of_arrival = timezone.now()

        order.status = status
        order.save()

        sms_url = 'https://webapp.usmsgh.com/api/sms/send'
        sms_message = f"Hello, {order.owner_name}. Your shipping order status has changed. Track your shipment with the link below.\nhhttps://www.ghbays.com/trackshipment/{tracking.tracking_number}"

        receiver_body = {
            'recipient': f"233{order.phone_number[1:]}",
            'sender_id': 'GH BAY',
            'message': sms_message
        }

        sms_headers = {
            'Authorization': 'Bearer 1334|wroIm5YnQD6hlZzd8POtLDXxl4vQodCZNorATYGX',
            'Content-Type': 'application/json'
        }

        response = requests.request('POST', url=sms_url, params=receiver_body,
                                    headers=sms_headers)
        print(response.text)
        # send_sms_notification(order.phone_number, order.status)
        return redirect('admin_order_list')

    orders = ShippingOrder.objects.all()
    return render(request, 'layouts/admin_orders.html', {'orders': orders})


from django.shortcuts import get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from .models import CustomUser

import string
import random


def generate_unique_shipping_code(length=8):
    characters = string.ascii_letters + string.digits
    return ''.join(random.choice(characters) for _ in range(length))


@login_required
def generate_shipping_code(request):
    user = request.user
    unique_code = generate_unique_shipping_code()

    # Ensure the code is unique
    while CustomUser.objects.filter(unique_shipping_code=unique_code).exists():
        unique_code = generate_unique_shipping_code()

    user.unique_shipping_code = unique_code
    user.save()
    messages.success(request, f"Your unique shipping code is {user.unique_shipping_code}. You can also find your code at the bottom or end of the page.")
    return redirect('home')
